#%% Oncobreast_env kernel
# Imports and function definitions
import tifffile
import numpy as np
import matplotlib.pyplot as plt
import random
import os
import pandas as pd
import datetime

from natsort import natsorted
from time import time
from cellpose import models
from tkinter.filedialog import askdirectory
from skimage.filters import gaussian, threshold_otsu, threshold_yen
from skimage.measure import regionprops_table
import openpyxl

def scramble(mask):
    idx = np.unique(mask[mask>0])
    idx_new = idx.copy()
    random.shuffle(idx)
    mask_new = np.zeros_like(mask)
    for n,i in enumerate(idx):
        mask_new[mask==idx[n]] = idx_new[n]
    return mask_new

#%% 
# Load Cellpose model and file selection
model_type='cyto3'
model = models.Cellpose(model_type=model_type,gpu = True)
DirMethod_all = askdirectory(title='Select Directory with Methods to be processed')
#%%
# Analysis parameters 
Cellpose_Diameter   = None              # Diameter for cells (in pixel, is scaled with bin_step)
                                        # None sets it automatically via Cellpose
BKG_sigma_gauss     = 20                # Sigma of the gaussian smoothing used to remove backgroud in fluorescence images
FLAG_SAME_CELLTYPE  = True              # if True, sets the Cellpose_Diameter to the one optimized in the first analyzed well
                                        # if False, optimizes for each well

Properties = [  'label', 
                'area', 
                'centroid', 
                'eccentricity',
                'solidity',
                'equivalent_diameter_area'
            ]
#%
# Loop through Methods in folder
for f in os.listdir(DirMethod_all):
    DirMethod = os.path.join(DirMethod_all,f)
    Analysis_path = DirMethod+'\Export\OncoBreast\\'
    try:
        Analysis_dir = os.mkdir(Analysis_path)
        print('Creating analysis directory at:'+Analysis_path)
    except:
        print('Analysis directory already exists at:'+Analysis_path)
    Filenames_images = [DirMethod+'/Images/'+f for f in os.listdir(DirMethod+'/Images')]

    # Columns specific for this type of analysis (time lapse, fluorescence + brightfield)
    Columns_filename = ['Acquisition Name',
                        'Well',
                        'ID Acquisition',
                        'ID Acquisition - Mode',
                        'Image Code and extension',
                        'Filename',
                        ]

    df = pd.DataFrame(columns=Columns_filename, index=[0])
    idx = 0
    for file in Filenames_images:
        if file.endswith('.tif') or file.endswith('.tiff'):
            file_base = os.path.basename(file)
            parts = file_base.split('_')

            df.loc[idx] = parts + [file]
            idx += 1
    #%
    # Main Loop - Well analysis
    FLAG_DF = True
    if FLAG_SAME_CELLTYPE:
        FLAG_TIME = False
    else:
        FLAG_TIME = True
    for acq in df['Acquisition Name'].unique():
        df_acq = df[df['Acquisition Name'] == acq]
        for well in natsorted(df_acq['Well'].unique()):
            df_tmp = df_acq[df_acq['Well'] == well]
            if FLAG_TIME:
                t0 = time()
            else:
                t0 = 0
            print(f'Processing {acq} - {well}...', end=' ')
            img_BF = tifffile.imread(df_tmp[df_tmp['ID Acquisition - Mode']=='Bf']['Filename'].iloc[0])  # Read the first file as a reference for BF
            x,y = img_BF.shape
            Npix = x*y
            masks, flows, styles, diams = model.eval(img_BF, diameter=Cellpose_Diameter, channels=[0,0])
            if FLAG_SAME_CELLTYPE:
                Cellpose_Diameter = diams
                print("Cellpose Diameter set to:"+str(Cellpose_Diameter))
                FLAG_SAME_CELLTYPE = False
                FLAG_TIME = True
            Cell_Number = np.max(masks)
            Confluency = np.sum(masks>0)/masks.size
            print(f'Found {Cell_Number} cells with confluency {Confluency:.2f}')
            df_exp = pd.DataFrame(columns=['Well','Cell Number','Confluency'],index = [0])
            df_exp.loc[0,'Well'] =  well  
            df_exp.loc[0,'Cell Number'] =  Cell_Number  
            df_exp.loc[0,'Confluency'] =  Confluency  
            df_exp.loc[0,'Cellpose Diameter'] =  diams  
            Acquisition_Mode_list = [mode for mode in df_tmp['ID Acquisition - Mode'].unique() if mode not in ['Ph', 'Bf']]
            if len(Acquisition_Mode_list) > 0:
                for Acquisition_Mode in Acquisition_Mode_list:
                    img_mod = tifffile.imread(df_tmp[df_tmp['ID Acquisition - Mode'] == Acquisition_Mode]['Filename'].iloc[0])
                    img_mod = img_mod - gaussian(img_mod, sigma=BKG_sigma_gauss,preserve_range=True)  # Apply Gaussian filter to the image
                    img_mod[img_mod < 0] = 0  # Ensure no negative values after filtering        
                    params = regionprops_table(masks, 
                                                intensity_image = img_mod, 
                                                properties=Properties+['intensity_mean','intensity_max'])
                    str_mode = ''
                    for s in ['ID Acquisition','ID Acquisition - Mode']:
                        str_mode = str_mode + df_tmp[df_tmp['ID Acquisition - Mode'] == Acquisition_Mode][s]
                    str_mode = str_mode.item()
                    params_df = pd.DataFrame(params)
                    columns_dict = {}
                    for c0 in params_df.columns[-2:]:
                        columns_dict[c0] = str_mode+' - '+ c0
                    params_df = params_df.rename(columns=columns_dict)
                    params_df['Acquisition Name'] = acq
                    params_df['Well'] = well
                for c in [c for c in params_df.columns if 'intensity' in c]:
                    df_exp.loc[0,'Npos Otsu-'+str_mode +'-'+ c.split('_')[-1]] = (params_df[c]>threshold_otsu(np.asarray(params_df[c]))).sum()
                    df_exp.loc[0,'Npos Yen-'+str_mode +'-'+ c.split('_')[-1]] = (params_df[c]>threshold_yen(np.asarray(params_df[c]))).sum()
            else:
                params = regionprops_table(masks, properties=Properties)
                params_df = pd.DataFrame(params)
                params_df['Well'] = well
                
            if FLAG_TIME:
                if t0 != 0:
                    t1 = time()
                    print(f'Time taken: {t1-t0:.2f} seconds')
                    t_estimated_time = (t1-t0) * (len(df_acq['Well'].unique()))
                    t_est_total = t_estimated_time * (len(df_acq['Well'].unique())) * (len(df['Acquisition Name'].unique()))
                    print(f'Estimated time for this acquisition: {t_estimated_time/60:.2f} minutes')
                    FLAG_TIME = False            
            
            if FLAG_DF:
                df_final = params_df.copy()
                df_exp_final = df_exp.copy()
                FLAG_DF = False
            else:
                df_final = pd.concat([df_final, params_df], ignore_index=True)    
                df_exp_final = pd.concat([df_exp_final, df_exp], ignore_index=True)    

    #%
    # Saving parameters and full single cell dataframe
    Params_Dict = {
        'Diameter': Cellpose_Diameter,
        'BKG_sigma_gauss': BKG_sigma_gauss,
        'Properties': Properties,
        'Columns_filename': Columns_filename,
        'Directory': DirMethod,
        'Analysis_path': Analysis_path,
    }
    df_final.to_csv(os.path.join(Analysis_path, 'Cellpose_Segmentation_Results.csv'), index=False)
    np.save(os.path.join(Analysis_path, 'Cellpose_Segmentation_Params.npy'), Params_Dict)
        
    # % 
    # Save Excel - Population level features
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws_cellnum = wb.active
    ws_cellnum.title = "Cell Number"

    # Define well rows and columns
    well_rows = list("ABCDEFGH")
    well_cols = [str(i) for i in range(1, 13)]

    # Add "Cell Number" sheet (matrix format)
    ws_cellnum.cell(row=1, column=1, value="")
    for col_idx, col in enumerate(well_cols, start=2):
        ws_cellnum.cell(row=1, column=col_idx, value=col)
    for row_idx, row in enumerate(well_rows, start=2):
        ws_cellnum.cell(row=row_idx, column=1, value=row)
        for col_idx, col in enumerate(well_cols, start=2):
            well_name = f"{row}{col}"
            cell_count = df_final[df_final['Well'] == well_name].shape[0]
            ws_cellnum.cell(row=row_idx, column=col_idx, value=cell_count)

    # Add "Confluency" sheet (matrix format)
    ws_confluency = wb.create_sheet(title="Confluency")
    ws_confluency.cell(row=1, column=1, value="")
    for col_idx, col in enumerate(well_cols, start=2):
        ws_cellnum.cell(row=1, column=col_idx, value=col)
    for row_idx, row in enumerate(well_rows, start=2):
        ws_cellnum.cell(row=row_idx, column=1, value=row)
        for col_idx, col in enumerate(well_cols, start=2):
            well_name = f"{row}{col}"
            mask_area = df_final[df_final['Well'] == well_name]['area'].sum()
            confluency = 100.0 * mask_area / Npix
            ws_confluency.cell(row=row_idx, column=col_idx, value=confluency)
            
    # For each property, create a new sheet and fill the 8x12 matrix
    properties_to_save = [col for col in df_final.columns if col not in ['centroid-0','centroid-1', 'label','Acquisition Name', 'Well']]
    for prop in properties_to_save:
        # Add a new sheet for each property except for the first (use the first as the main sheet)
        ws = wb.create_sheet(title=prop)
        # Write headers
        ws.cell(row=1, column=1, value="")
        for col_idx, col in enumerate(well_cols, start=2):
            ws.cell(row=1, column=col_idx, value=col)
        for row_idx, row in enumerate(well_rows, start=2):
            ws.cell(row=row_idx, column=1, value=row)

        # Fill the matrix with medians
        for row_idx, row in enumerate(well_rows, start=2):
            for col_idx, col in enumerate(well_cols, start=2):
                well_name = f"{row}{col}"
                median_val = df_final[df_final['Well'] == well_name][prop].median()
                ws.cell(row=row_idx, column=col_idx, value=median_val)
    # For each property, create a new sheet and fill the 8x12 matrix
    properties_to_save = [col for col in df_exp_final.columns if col not in ['Cell Number','Confluency','Well']]
    for prop in properties_to_save:
        # Add a new sheet for each property except for the first (use the first as the main sheet)
        ws = wb.create_sheet(title=prop)
        # Write headers
        ws.cell(row=1, column=1, value="")
        for col_idx, col in enumerate(well_cols, start=2):
            ws.cell(row=1, column=col_idx, value=col)
        for row_idx, row in enumerate(well_rows, start=2):
            ws.cell(row=row_idx, column=1, value=row)

        # Fill the matrix with values
        for row_idx, row in enumerate(well_rows, start=2):
            for col_idx, col in enumerate(well_cols, start=2):
                well_name = f"{row}{col}"
                median_val = df_exp_final[df_exp_final['Well'] == well_name][prop]
                if len(median_val) == 0:
                    median_val = 0
                else:
                    median_val = median_val.values[0]
                ws.cell(row=row_idx, column=col_idx, value=median_val)

    # Save the workbook
    xlsx_path = os.path.join(Analysis_path, "Cellpose_Segmentation_Medians.xlsx")
    wb.save(xlsx_path)
    print(f"Median spreadsheet saved to {xlsx_path}")

    # %
    # Save Excel - Histograms
    Nbins_histogram = 32
    CMap = 'nipy_spectral'

    # Create a new workbook for histograms
    wb_hist = openpyxl.Workbook()
    ws_hist = wb_hist.active
    ws_hist.title = "Histograms"
    # Define well order: A1, A2, ..., H12 (column increases first)
    well_order = natsorted([f"{row}{col}" for col in well_cols for row in well_rows])

    import matplotlib.pyplot as plt

    # Exclude 'centroid' and 'label' properties for histograms
    histogram_props = [col for col in df_final.columns if col not in ['centroid-0','centroid-1', 'label','Acquisition Name', 'Well']]

    for prop in histogram_props:
        ws_prop = wb_hist.create_sheet(title=prop)
        ws_prop.cell(row=1, column=1, value="Well")
        bin_edges = None
        histograms = []
        all_data = df_final[prop].dropna()
        if len(all_data) == 0:
            min_val, max_val = 0, 1
        else:
            min_val, max_val = float(np.percentile(all_data,1)), float(np.percentile(all_data,99))
            if min_val == max_val:
                max_val = min_val + 1e-6  # avoid zero-width bins

        # Use log scale for 'area' and intensity properties
        is_log = False
        if prop == 'area' or 'intensity' in prop:
            is_log = True

        if is_log:
            # Avoid log(0) by setting minimum > 0
            min_val = max(min_val, 1e-6)
            bins = np.logspace(np.log10(min_val), np.log10(max_val), Nbins_histogram+1)
        else:
            bins = np.linspace(min_val, max_val, Nbins_histogram+1)

        # Compute histograms for each well
        for idx, well in enumerate(well_order, start=2):
            data = df_final[df_final['Well'] == well][prop].dropna()
            if is_log:
                data = data[data > 0]  # Remove non-positive values for log scale
            if len(data) == 0:
                hist = [0]*Nbins_histogram
            else:
                hist, _ = np.histogram(data, bins=bins)
            histograms.append(hist)
            ws_prop.cell(row=idx, column=1, value=well)
            for j, val in enumerate(hist, start=2):
                ws_prop.cell(row=idx, column=j, value=int(val))

        # Write bin edges after the last well
        ws_prop.cell(row=len(well_order)+2, column=1, value="Bin edges")
        for j, edge in enumerate(bins, start=2):
            ws_prop.cell(row=len(well_order)+2, column=j, value=float(edge))

        # Plot all histograms with legend, excluding all-zero histograms
        plt.figure(figsize=(12, 6))
        cmap = plt.get_cmap(CMap)
        color_idx = 0
        n_hist = 0
        for hist in histograms:
            if np.sum(hist) != 0:
                n_hist += 1
        for idx, hist in enumerate(histograms):
            if np.sum(hist) == 0:
                continue  # Skip all-zero histograms
            color = cmap(color_idx / max(1, n_hist-1))
            if is_log:
                plt.semilogx(bins[:-1], hist, label=well_order[idx], color=color)
            else:
                plt.plot(bins[:-1], hist, label=well_order[idx], color=color)
            color_idx += 1
        plt.xlabel(prop)
        plt.ylabel("Count")
        plt.title(f"Histograms for {prop}")
        plt.legend(fontsize='x-small', ncol=4)
        plt.tight_layout()
        plot_path = os.path.join(Analysis_path, f"Histogram_{prop}.png")
        plt.savefig(plot_path)
        plt.close()

    # Remove the default sheet if it exists
    if "Histograms" in wb_hist.sheetnames:
        del wb_hist["Histograms"]

    xlsx_hist_path = os.path.join(Analysis_path, "Cellpose_Segmentation_Histograms.xlsx")
    wb_hist.save(xlsx_hist_path)
    print(f"Histogram spreadsheet saved to {xlsx_hist_path}")

    # %%
